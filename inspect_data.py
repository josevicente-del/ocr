import openpyxl
import fitz

def inspect_excel():
    print("=== INSPECTING EXCEL ===")
    wb = openpyxl.load_workbook("Cosas/ejemplp.xlsx")
    print("Sheets:", wb.sheetnames)
    ws = wb.active
    print("Dimensions:", ws.dimensions)
    for r in range(1, 15):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, 15)]
        if any(row_vals):
            print(f"Row {r}: {row_vals}")

def inspect_pdf():
    print("\n=== INSPECTING PDF ===")
    doc = fitz.open("Cosas/20260604142405.pdf")
    print("Page count:", len(doc))
    
    # Let's inspect page 1 and page 5 details
    for page_idx in [0, 4, 10]:
        if page_idx >= len(doc):
            break
        page = doc[page_idx]
        print(f"\n--- PAGE {page_idx+1} ---")
        text = page.get_text("text")
        print("TEXT LENGTH:", len(text))
        print("FIRST 1000 CHARS OF TEXT:")
        print(text[:1000])
        
        # Let's print drawing objects or block details if any
        blocks = page.get_text("blocks")
        print("BLOCKS COUNT:", len(blocks))
        for idx, b in enumerate(blocks[:10]):
            print(f"Block {idx}: Rect={b[:4]}, Text={repr(b[4])}")

if __name__ == "__main__":
    inspect_excel()
    inspect_pdf()
