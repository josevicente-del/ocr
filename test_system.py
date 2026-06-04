# -*- coding: utf-8 -*-
"""
Pruebas de Integración y API del Sistema.
Valida todos los endpoints de FastAPI utilizando TestClient.
"""

import os
import json
import openpyxl
from fastapi.testclient import TestClient

# Asegurar que se importan correctamente la app y el estado
from app import app, db_state, DB_FILE

client = TestClient(app)

def test_api_endpoints():
    print("=== INICIANDO PRUEBAS DE API ===")
    
    # 1. Probar el reinicio (Reset)
    print("Probando POST /api/reset...")
    response = client.post("/api/reset")
    assert response.status_code == 200
    assert response.json()["message"] == "Aplicación reiniciada correctamente."
    assert len(db_state["orders"]) == 0
    print("  -> OK")
    
    # 2. Probar el estado inicial (Status)
    print("Probando GET /api/status inicial...")
    response = client.get("/api/status")
    assert response.status_code == 200
    data = response.json()
    assert data["is_processing"] is False
    assert data["total_pages"] == 0
    assert data["articles_count"] == 0
    assert data["unresolved_count"] == 0
    print("  -> OK")
    
    # 3. Probar configuración de columnas (Columns)
    print("Probando POST /api/columns...")
    custom_order = [
        "Nmero Pedido Cliente",
        "Fecha Pedido",
        "Cliente",
        "BAR Pedidas",
        "DESDESG",
        "Tratamiento"
    ]
    response = client.post("/api/columns", json={"column_order": custom_order})
    assert response.status_code == 200
    assert db_state["column_order"] == custom_order
    print("  -> OK")
    
    # 4. Inyectar datos simulados de pedido para probar la resolución y exportación
    print("Inyectando datos de prueba...")
    db_state["orders"] = {
        "1": {
            "order_number": "38-2026F0110003587",
            "date": "04/06/2026",
            "client": "ALUMELUM EN MURCIA",
            "articles": [
                {
                    "code": "08255964",
                    "quantity": 2,
                    "description": "MARCO SUPERIOR",
                    "treatment_raw": "RAL BLANCO",
                    "treatment_mapped": "PENDIENTE_CONFIRMACION",
                    "needs_resolution": True
                }
            ],
            "had_problem": True,
            "problem_details": ["Fila 1: Tratamiento desconocido 'RAL BLANCO'"]
        }
    }
    
    # 5. Probar estado con datos no resueltos
    print("Probando GET /api/status con datos no resueltos...")
    response = client.get("/api/status")
    assert response.status_code == 200
    data = response.json()
    assert data["unresolved_count"] == 1
    assert data["unresolved_items"][0]["treatment_raw"] == "RAL BLANCO"
    print("  -> OK")
    
    # 6. Probar resolución de tratamientos (Resolve)
    print("Probando POST /api/resolve...")
    response = client.post("/api/resolve", json={
        "treatment_raw": "RAL BLANCO",
        "treatment_mapped": "RAL BLANCO ERP"
    })
    assert response.status_code == 200
    assert response.json()["updated_count"] == 1
    
    # Verificar que el estado se actualizó
    response = client.get("/api/status")
    data = response.json()
    assert data["unresolved_count"] == 0
    assert data["orders_data"]["1"]["articles"][0]["treatment_mapped"] == "RAL BLANCO ERP"
    assert data["orders_data"]["1"]["articles"][0]["needs_resolution"] is False
    print("  -> OK")
    
    # 7. Probar exportación a Excel (Export)
    print("Probando GET /api/export...")
    response = client.get("/api/export")
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    
    # Escribir el archivo temporal y validarlo con openpyxl
    temp_file = "temp_test_export.xlsx"
    with open(temp_file, "wb") as f:
        f.write(response.content)
        
    wb = openpyxl.load_workbook(temp_file)
    assert "Hoja1" in wb.sheetnames
    ws = wb["Hoja1"]
    
    # Verificar cabeceras de columnas configuradas
    headers = [ws.cell(row=1, column=c).value for c in range(1, 7)]
    assert headers == custom_order
    
    # Verificar datos de fila
    row2 = [ws.cell(row=2, column=c).value for c in range(1, 7)]
    assert row2[0] == "38-2026F0110003587"  # Nmero Pedido Cliente
    assert row2[1] == "04/06/2026"           # Fecha Pedido
    assert row2[2] == "ALUMELUM EN MURCIA"   # Cliente
    assert row2[3] == 2                      # BAR Pedidas
    assert row2[4] == "08255964 MARCO SUPERIOR" # DESDESG
    assert row2[5] == "RAL BLANCO ERP"       # Tratamiento
    
    wb.close()
    os.remove(temp_file)
    print("  -> OK")
    
    # Restaurar base de datos inicial y limpiar db.json de prueba
    client.post("/api/reset")
    if os.path.exists(DB_FILE):
        os.remove(DB_FILE)
        
    print("\n=== ¡TODAS LAS PRUEBAS DE API PASARON EXITOSAMENTE! ===")

if __name__ == "__main__":
    test_api_endpoints()
